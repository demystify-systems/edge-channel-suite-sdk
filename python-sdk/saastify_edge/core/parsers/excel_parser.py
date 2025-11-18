"""
Excel (XLSX/XLSM) file parser with streaming support
"""

from typing import AsyncIterator, Dict, Any, Optional
from .base import BaseParser
from ..types import FileConfig

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelParser(BaseParser):
    """Parser for Excel files (XLSX, XLSM)"""
    
    def __init__(self, file_config: Optional[FileConfig] = None):
        super().__init__(file_config)
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
    
    def can_parse(self, file_path: str) -> bool:
        """Check if file is Excel"""
        return file_path.lower().endswith(('.xlsx', '.xlsm', '.xls'))
    
    async def parse(self, file_path: str) -> AsyncIterator[Dict[str, Any]]:
        """
        Parse Excel file and yield rows.
        
        Uses read_only mode for memory efficiency with large files.
        """
        # Load workbook in read-only mode for streaming
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Select worksheet
        if self.sheet_name:
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")
        else:
            # Use first sheet
            ws = wb.active
        
        # Iterate through rows
        headers = None
        row_num = 0
        
        for excel_row in ws.iter_rows(values_only=True):
            row_num += 1
            
            # Skip fixed rows
            if row_num <= self.fixed_rows:
                continue
            
            # Extract headers
            if row_num == self.header_row:
                headers = [
                    str(col).strip() if col is not None else f"Column_{i}"
                    for i, col in enumerate(excel_row)
                ]
                continue
            
            # Skip if headers not yet read
            if headers is None:
                continue
            
            # Build row dictionary
            row_data = {}
            for i, value in enumerate(excel_row):
                col_name = headers[i] if i < len(headers) else f"Column_{i}"
                
                # Convert value to string, handle None
                if value is None:
                    row_data[col_name] = None
                elif isinstance(value, (int, float)):
                    row_data[col_name] = value
                else:
                    row_data[col_name] = str(value).strip()
            
            yield {
                "file_row_number": row_num,
                "data": row_data,
                "raw_input_snapshot": row_data.copy()
            }
        
        wb.close()
