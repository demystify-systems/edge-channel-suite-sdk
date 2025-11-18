"""
File Builders - Generate export files in various formats.

Supports:
- CSV
- TSV
- XLSX (Excel)
- JSON
- XML
"""

import csv
import json
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class FileBuilderError(Exception):
    """Base exception for file building errors."""
    pass


class FileBuilder:
    """Base class for file builders."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build file from data.
        
        Args:
            data: List of row dictionaries
            output_path: Path to output file
            config: Optional configuration
            
        Returns:
            Path to created file
        """
        raise NotImplementedError


class CSVFileBuilder(FileBuilder):
    """Build CSV files."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build CSV file.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            config: Optional config (delimiter, headers, quoting)
            
        Returns:
            Path to created file
        """
        config = config or {}
        delimiter = config.get("delimiter", ",")
        include_headers = config.get("include_headers", True)
        quoting = config.get("quoting", csv.QUOTE_MINIMAL)

        if not data:
            raise FileBuilderError("No data to export")

        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                # Get all unique field names
                fieldnames = list(data[0].keys()) if data else []

                writer = csv.DictWriter(
                    f,
                    fieldnames=fieldnames,
                    delimiter=delimiter,
                    quoting=quoting,
                )

                if include_headers:
                    writer.writeheader()

                writer.writerows(data)

            logger.info(f"Created CSV file with {len(data)} rows: {output_path}")
            return output_path

        except Exception as e:
            raise FileBuilderError(f"CSV file creation failed: {e}")


class TSVFileBuilder(FileBuilder):
    """Build TSV files."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build TSV file.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            config: Optional config (headers, quoting)
            
        Returns:
            Path to created file
        """
        # TSV is just CSV with tab delimiter
        config = config or {}
        config["delimiter"] = "\t"
        
        csv_builder = CSVFileBuilder()
        return csv_builder.build_file(data, output_path, config)


class XLSXFileBuilder(FileBuilder):
    """Build Excel (XLSX) files."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build XLSX file.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            config: Optional config (sheet_name, include_headers)
            
        Returns:
            Path to created file
        """
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            raise FileBuilderError(
                "openpyxl not installed. Install with: pip install openpyxl"
            )

        config = config or {}
        sheet_name = config.get("sheet_name", "Products")
        include_headers = config.get("include_headers", True)

        if not data:
            raise FileBuilderError("No data to export")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Get field names
            fieldnames = list(data[0].keys())

            # Write headers
            if include_headers:
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    ws.cell(row=1, column=col_idx, value=fieldname)
                start_row = 2
            else:
                start_row = 1

            # Write data rows
            for row_idx, row_data in enumerate(data, start=start_row):
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    value = row_data.get(fieldname)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(output_path)
            logger.info(f"Created XLSX file with {len(data)} rows: {output_path}")
            return output_path

        except Exception as e:
            raise FileBuilderError(f"XLSX file creation failed: {e}")


class JSONFileBuilder(FileBuilder):
    """Build JSON files."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build JSON file.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            config: Optional config (indent, format='array' or 'ndjson')
            
        Returns:
            Path to created file
        """
        config = config or {}
        indent = config.get("indent", 2)
        json_format = config.get("format", "array")  # 'array' or 'ndjson'

        if not data:
            raise FileBuilderError("No data to export")

        try:
            with open(output_path, "w", encoding="utf-8") as f:
                if json_format == "ndjson":
                    # Newline-delimited JSON
                    for row in data:
                        f.write(json.dumps(row, ensure_ascii=False) + "\n")
                else:
                    # Standard JSON array
                    json.dump(data, f, indent=indent, ensure_ascii=False)

            logger.info(f"Created JSON file with {len(data)} rows: {output_path}")
            return output_path

        except Exception as e:
            raise FileBuilderError(f"JSON file creation failed: {e}")


class XMLFileBuilder(FileBuilder):
    """Build XML files."""

    def build_file(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Build XML file.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            config: Optional config (root_tag, row_tag)
            
        Returns:
            Path to created file
        """
        try:
            import xml.etree.ElementTree as ET
        except ImportError:
            raise FileBuilderError("xml.etree.ElementTree not available")

        config = config or {}
        root_tag = config.get("root_tag", "products")
        row_tag = config.get("row_tag", "product")

        if not data:
            raise FileBuilderError("No data to export")

        try:
            # Create root element
            root = ET.Element(root_tag)

            # Add data rows
            for row_data in data:
                row_element = ET.SubElement(root, row_tag)
                
                for key, value in row_data.items():
                    field_element = ET.SubElement(row_element, str(key))
                    field_element.text = str(value) if value is not None else ""

            # Create tree and write
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ")  # Pretty print (Python 3.9+)
            tree.write(output_path, encoding="utf-8", xml_declaration=True)

            logger.info(f"Created XML file with {len(data)} rows: {output_path}")
            return output_path

        except Exception as e:
            raise FileBuilderError(f"XML file creation failed: {e}")


class FileBuilderFactory:
    """Factory to create appropriate file builder based on format."""

    @staticmethod
    def create_builder(file_format: str) -> FileBuilder:
        """
        Create appropriate file builder for format.
        
        Args:
            file_format: File format (csv, tsv, xlsx, json, xml)
            
        Returns:
            FileBuilder instance
        """
        file_format = file_format.lower()

        if file_format == "csv":
            return CSVFileBuilder()
        elif file_format == "tsv":
            return TSVFileBuilder()
        elif file_format in ["xlsx", "excel"]:
            return XLSXFileBuilder()
        elif file_format == "json":
            return JSONFileBuilder()
        elif file_format == "xml":
            return XMLFileBuilder()
        else:
            raise FileBuilderError(f"Unsupported file format: {file_format}")

    @staticmethod
    def build_file(
        data: List[Dict[str, Any]],
        output_path: str,
        file_format: Optional[str] = None,
        config: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Convenience method to build file with automatic format detection.
        
        Args:
            data: List of row dictionaries
            output_path: Output file path
            file_format: File format (if None, inferred from output_path extension)
            config: Optional configuration
            
        Returns:
            Path to created file
        """
        if not file_format:
            # Infer from file extension
            ext = Path(output_path).suffix.lstrip(".")
            file_format = ext if ext else "csv"

        builder = FileBuilderFactory.create_builder(file_format)
        return builder.build_file(data, output_path, config)
