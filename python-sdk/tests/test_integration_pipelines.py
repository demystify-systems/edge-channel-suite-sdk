"""
Integration tests for import and export pipelines.

Tests end-to-end workflows:
- Import: file fetch → parse → transform → validate → cache
- Export: fetch products → transform → validate → cache → build file
"""

import asyncio
import os
import tempfile
from pathlib import Path

try:
    import pytest
    HAS_PYTEST = True
except ImportError:
    HAS_PYTEST = False
    # Mock pytest.mark.asyncio decorator
    class MockPytest:
        class mark:
            @staticmethod
            def asyncio(func):
                return func
    pytest = MockPytest()

from saastify_edge.import_pipeline.orchestrator import (
    ImportPipeline,
    ImportPipelineConfig,
    run_product_import,
)
from saastify_edge.export.orchestrator import (
    ExportPipeline,
    ExportPipelineConfig,
    run_product_export,
)
from saastify_edge.db.mock_db_client import MockDBClient


@pytest.mark.asyncio
async def test_import_pipeline_csv():
    """Test full import pipeline with CSV file."""
    # Create test CSV file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("product_title,selling_price,stock_qty,product_sku,product_description\n")
        f.write("wireless headphones,$49.99,150,WH-001,Premium audio\n")
        f.write("smart watch,99.00,75,SW-002,Fitness tracker\n")
        f.write("power bank,29.99,200,PB-003,20000mAh capacity\n")
        csv_file = f.name

    try:
        # Create mock database
        mock_db = MockDBClient()
        
        # Configure import
        config = ImportPipelineConfig(
            file_source=csv_file,
            template_id="test-template-001",
            saas_edge_id="test-tenant-001",
            job_name="Test Import",
            batch_size=2,
            max_workers=2,
            db_client=mock_db,
        )

        # Run import pipeline
        results = await run_product_import(config)

        # Verify results
        assert results["total_processed"] == 3
        assert results["total_errors"] == 0
        assert results["total_batches"] == 2  # 3 rows in batches of 2
        assert results["duration_seconds"] > 0
        
        print(f"✅ Import pipeline processed {results['total_processed']} rows successfully")

    finally:
        os.unlink(csv_file)


@pytest.mark.asyncio
async def test_export_pipeline_csv():
    """Test full export pipeline generating CSV."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = os.path.join(tmpdir, "export_test.csv")

        # Configure export
        config = ExportPipelineConfig(
            template_id="test-template-001",
            saas_edge_id="test-tenant-001",
            output_path=output_file,
            file_format="csv",
            job_name="Test Export CSV",
            db_client=MockDBClient(),
        )

        # Run export pipeline
        results = await run_product_export(config)

        # Verify results
        assert results["total_rows"] == 3  # Mock data has 3 products
        assert os.path.exists(results["output_path"])
        
        # Verify file content
        with open(output_file, "r") as f:
            lines = f.readlines()
            assert len(lines) == 4  # 1 header + 3 rows
            assert "title" in lines[0]
            assert "Wireless Bluetooth Headphones" in lines[1]  # Transformed (title_case)
        
        print(f"✅ Export pipeline generated file with {results['total_rows']} rows")


@pytest.mark.asyncio
async def test_export_pipeline_json():
    """Test export pipeline generating JSON."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = os.path.join(tmpdir, "export_test.json")

        config = ExportPipelineConfig(
            template_id="test-template-001",
            saas_edge_id="test-tenant-001",
            output_path=output_file,
            file_format="json",
            job_name="Test Export JSON",
            db_client=MockDBClient(),
        )

        results = await run_product_export(config)

        assert results["total_rows"] == 3
        assert os.path.exists(output_file)
        
        # Verify JSON format
        import json
        with open(output_file, "r") as f:
            data = json.load(f)
            assert isinstance(data, list)
            assert len(data) == 3
            assert "title" in data[0]
        
        print("✅ Export pipeline generated JSON successfully")


@pytest.mark.asyncio
async def test_export_pipeline_xlsx():
    """Test export pipeline generating Excel file."""
    if HAS_PYTEST:
        pytest.importorskip("openpyxl")
    else:
        try:
            import openpyxl
        except ImportError:
            print("⚠️  Skipping XLSX test - openpyxl not installed")
            return
    
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = os.path.join(tmpdir, "export_test.xlsx")

        config = ExportPipelineConfig(
            template_id="test-template-001",
            saas_edge_id="test-tenant-001",
            output_path=output_file,
            file_format="xlsx",
            job_name="Test Export XLSX",
            file_config={"sheet_name": "Products"},
            db_client=MockDBClient(),
        )

        results = await run_product_export(config)

        assert results["total_rows"] == 3
        assert os.path.exists(output_file)
        
        # Verify Excel file
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active
        assert ws.title == "Products"
        assert ws.max_row == 4  # 1 header + 3 rows
        
        print("✅ Export pipeline generated XLSX successfully")


@pytest.mark.asyncio
async def test_template_mapper():
    """Test template loading and field mapping."""
    from saastify_edge.import_pipeline.template_mapper import TemplateMapper
    
    mapper = TemplateMapper()
    
    # Load template
    template = await mapper.load_template(
        template_id="test-template-001",
        saas_edge_id="test-tenant-001",
    )
    
    assert template.id == "test-template-001"
    assert template.channel_name == "amazon"
    assert len(template.attributes) > 0
    
    # Test field mapping
    raw_row = {
        "product_title": "Test Product",
        "selling_price": "19.99",
        "stock_qty": "100",
        "product_sku": "TEST-001",
    }
    
    mapped = mapper.map_row_to_fields(raw_row, template)
    
    assert "title" in mapped
    assert "price" in mapped
    assert mapped["title"] == "Test Product"
    
    print("✅ Template mapper working correctly")


@pytest.mark.asyncio
async def test_batch_processor():
    """Test batch processing with concurrency."""
    from saastify_edge.import_pipeline.batch_processor import BatchProcessor, BatchConfig
    
    async def mock_data_stream():
        """Generate test data."""
        for i in range(10):
            yield {"id": i, "value": f"item-{i}"}
    
    processed = []
    
    async def process_batch(batch):
        """Mock batch processor."""
        processed.extend(batch)
        return {
            "success_count": len(batch),
            "error_count": 0,
            "errors": [],
        }
    
    config = BatchConfig(batch_size=3, max_workers=2)
    processor = BatchProcessor(config)
    
    results = await processor.process_stream(
        data_stream=mock_data_stream(),
        processor_func=process_batch,
    )
    
    assert results["total_processed"] == 10
    assert results["total_errors"] == 0
    assert len(processed) == 10
    
    print(f"✅ Batch processor handled {results['total_processed']} rows")


@pytest.mark.asyncio
async def test_file_loaders():
    """Test file loaders for different sources."""
    from saastify_edge.core.loaders import LocalFileLoader
    
    # Create test file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        f.write("test content")
        test_file = f.name
    
    try:
        loader = LocalFileLoader()
        loaded_path = await loader.load_file(test_file)
        
        assert os.path.exists(loaded_path)
        with open(loaded_path) as f:
            content = f.read()
            assert content == "test content"
        
        print("✅ File loader working correctly")
        
    finally:
        os.unlink(test_file)


@pytest.mark.asyncio
async def test_file_builders():
    """Test file builders for different formats."""
    from saastify_edge.export.file_builders import (
        CSVFileBuilder,
        JSONFileBuilder,
    )
    
    test_data = [
        {"name": "Product A", "price": 10.99},
        {"name": "Product B", "price": 20.99},
    ]
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Test CSV builder
        csv_path = os.path.join(tmpdir, "test.csv")
        csv_builder = CSVFileBuilder()
        csv_builder.build_file(test_data, csv_path)
        assert os.path.exists(csv_path)
        
        # Test JSON builder
        json_path = os.path.join(tmpdir, "test.json")
        json_builder = JSONFileBuilder()
        json_builder.build_file(test_data, json_path)
        assert os.path.exists(json_path)
        
        print("✅ File builders working correctly")


def test_transformation_pipeline_integration():
    """Test transformation pipeline with mock template."""
    from saastify_edge.transformations.engine import apply_transformations
    
    # Test chained transformations
    result = apply_transformations(
        value="  hello world  ",
        rule_string="strip + uppercase + replace| |_"
    )
    assert result == "HELLO_WORLD"
    
    # Test numeric transformations
    result = apply_transformations(
        value="$49.99",
        rule_string="clean_numeric_value"
    )
    assert result == 49.99
    
    print("✅ Transformation pipeline integration working")


def test_validation_pipeline_integration():
    """Test validation pipeline with mock template."""
    from saastify_edge.validation.engine import validate_field
    
    # Test required validation
    error = validate_field(value=None, rule_name="required")
    assert error is not None
    
    # Test max_length validation
    error = validate_field(value="test", rule_name="max_length", max=10)
    assert error is None
    
    error = validate_field(value="a" * 100, rule_name="max_length", max=10)
    assert error is not None
    
    print("✅ Validation pipeline integration working")


if __name__ == "__main__":
    print("\n" + "="*60)
    print("Running Integration Tests")
    print("="*60 + "\n")
    
    # Run tests
    asyncio.run(test_import_pipeline_csv())
    asyncio.run(test_export_pipeline_csv())
    asyncio.run(test_export_pipeline_json())
    asyncio.run(test_export_pipeline_xlsx())
    asyncio.run(test_template_mapper())
    asyncio.run(test_batch_processor())
    asyncio.run(test_file_loaders())
    asyncio.run(test_file_builders())
    test_transformation_pipeline_integration()
    test_validation_pipeline_integration()
    
    print("\n" + "="*60)
    print("✅ ALL INTEGRATION TESTS PASSED")
    print("="*60)
